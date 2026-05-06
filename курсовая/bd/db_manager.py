# bd/db_manager.py
import os
from openpyxl import Workbook, load_workbook
from .data_init import DB_FILE, INITIAL_CARS


def init_excel_file():
    """Создает Excel файл, если его нет."""
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Cars"

        headers = ["Название", "Марка авто", "Мощность (л.с.)",
                   "Гарантия (пробег, тыс.км)", "Гарантия (лет)",
                   "Клиренс (мм)", "Стоимость (тыс.руб.)"]
        ws.append(headers)

        for row in INITIAL_CARS:
            ws.append(row)

        wb.save(DB_FILE)
        return True
    return False


def load_data_from_excel():
    """Загружает данные из Excel в словарь. Если файла нет, создает его."""
    if not os.path.exists(DB_FILE):
        init_excel_file()

    cars = {}
    try:
        wb = load_workbook(DB_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue

            name = str(row[0])
            try:
                maker = str(row[1])
                power = float(row[2])
                guar_km = float(row[3])
                guar_years = float(row[4])
                clearance = float(row[5])
                price = float(row[6])

                cars[name] = [maker, power, guar_km, guar_years, clearance, price]
            except (ValueError, TypeError, IndexError):
                continue
        wb.close()
    except Exception as e:
        print(f"Ошибка чтения файла {DB_FILE}: {e}")
        return {}

    return cars


def save_data_to_excel(cars):
    """Сохраняет словарь cars в Excel файл."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cars"

    headers = ["Название", "Марка авто", "Мощность (л.с.)",
               "Гарантия (пробег, тыс.км)", "Гарантия (лет)",
               "Клиренс (мм)", "Стоимость (тыс.руб.)"]
    ws.append(headers)

    for name, vals in cars.items():
        ws.append([name, vals[0], vals[1], vals[2], vals[3], vals[4], vals[5]])

    wb.save(DB_FILE)
